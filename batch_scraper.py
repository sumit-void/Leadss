"""
Batch Google Maps Scraper — Local Runner
  - Reads queries from queries.txt (one per line)
  - Runs each query using the same scraper logic
  - Combines ALL results into one Excel file (one sheet per query)
  - Also creates a combined "All Results" sheet
  - Skips businesses that already have a website

Usage:
    python batch_scraper.py
    python batch_scraper.py --max 40          # max results per query (default 60)
    python batch_scraper.py --file my_queries.txt  # custom queries file
    python batch_scraper.py --headless        # run without visible browser
"""

import re
import os
import sys
import time
import argparse
from datetime import datetime
from urllib.parse import quote_plus

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ai_enrichment import enrich_records


# ═══════════════════════════════════════════════════════════════
#  SCRAPER CORE (same logic as scrape_google_maps.py)
# ═══════════════════════════════════════════════════════════════

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()[:80]


def scroll_results(page, max_results: int = 60):
    feed = page.query_selector('div[role="feed"]')
    if not feed:
        return 0

    last_count = 0
    stale = 0

    while stale < 10:
        items = page.query_selector_all('div[role="feed"] > div > div > a')
        count = len(items)

        if count >= max_results:
            break

        if count == last_count:
            stale += 1
        else:
            stale = 0
            last_count = count

        feed.evaluate("el => el.scrollTop = el.scrollHeight")
        time.sleep(1.8)

    return min(len(page.query_selector_all('div[role="feed"] > div > div > a')), max_results)


def collect_place_urls(page, max_results: int = 60) -> list[str]:
    links = page.query_selector_all('div[role="feed"] > div > div > a')
    urls = []
    for link in links[:max_results]:
        href = link.get_attribute("href")
        if href and "/maps/place/" in href:
            urls.append(href)
    return urls


def extract_from_place_page(page, url: str, index: int) -> dict | None:
    try:
        page.goto(url, wait_until="load", timeout=30000)
        time.sleep(3)

        try:
            page.wait_for_selector('h1.DUwDvf, h1.fontHeadlineLarge', timeout=10000)
        except PWTimeout:
            try:
                page.wait_for_selector('h1', timeout=5000)
            except PWTimeout:
                return None

        # Scroll detail panel to reveal all info
        detail_panels = page.query_selector_all('div[role="main"]')
        for panel in detail_panels:
            try:
                for _ in range(5):
                    panel.evaluate("el => el.scrollTop += 500")
                    time.sleep(0.3)
                panel.evaluate("el => el.scrollTop = 0")
                time.sleep(0.3)
            except Exception:
                pass

        try:
            page.evaluate("""
                const scrollable = document.querySelector('[class*="m6QErb"][class*="DxyBCb"]');
                if (scrollable) {
                    for (let i = 0; i < 6; i++) {
                        scrollable.scrollTop += 400;
                    }
                    scrollable.scrollTop = 0;
                }
            """)
            time.sleep(0.5)
        except Exception:
            pass

        data = {
            "sr_no": index,
            "name": "",
            "address": "",
            "phone": "",
            "email": "",
            "website": "",
            "rating": "",
            "total_reviews": "",
            "category": "",
        }

        # Name
        for sel in ['h1.DUwDvf', 'h1.fontHeadlineLarge', 'h1']:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                if txt and txt.lower() != "results":
                    data["name"] = txt
                    break

        if not data["name"]:
            return None

        # Rating
        el = page.query_selector('div.F7nice span[aria-hidden="true"]')
        if el:
            data["rating"] = el.inner_text().strip()

        # Reviews
        el = page.query_selector('div.F7nice span[aria-label*="review"]')
        if not el:
            el = page.query_selector('div.F7nice span:nth-child(2)')
        if el:
            data["total_reviews"] = el.inner_text().strip().replace("(", "").replace(")", "").replace(",", "")

        # Category
        el = page.query_selector('button[jsaction*="category"]')
        if el:
            data["category"] = el.inner_text().strip()

        # Address
        for sel in ['button[data-item-id="address"]', 'button[data-tooltip="Copy address"]']:
            el = page.query_selector(sel)
            if el:
                data["address"] = el.inner_text().strip()
                break

        # Phone
        for sel in ['button[data-item-id*="phone:tel"]', 'button[data-tooltip="Copy phone number"]']:
            el = page.query_selector(sel)
            if el:
                data["phone"] = el.inner_text().strip()
                break

        # Email — 3-layer extraction
        email_selectors = [
            'a[data-item-id*="email"]',
            'button[data-item-id*="email"]',
            'a[href^="mailto:"]',
            'a[data-tooltip="Send email"]',
            'button[data-tooltip="Send email"]',
            '[data-item-id*="email"]',
            'a[aria-label*="email"]',
            'a[aria-label*="Email"]',
            'button[aria-label*="email"]',
            'button[aria-label*="Email"]',
        ]
        for sel in email_selectors:
            el = page.query_selector(sel)
            if el:
                href = el.get_attribute("href") or ""
                if href.startswith("mailto:"):
                    data["email"] = href.replace("mailto:", "").split("?")[0].strip()
                else:
                    aria = el.get_attribute("aria-label") or ""
                    txt = el.inner_text().strip()
                    email_in_aria = re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', aria)
                    if email_in_aria:
                        data["email"] = email_in_aria.group()
                    elif "@" in txt:
                        data["email"] = txt
                if data["email"]:
                    break

        if not data["email"]:
            try:
                info_elements = page.query_selector_all('div[role="main"] button, div[role="main"] a')
                for elem in info_elements:
                    txt = elem.inner_text().strip()
                    if "@" in txt and "." in txt:
                        match = re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', txt)
                        if match:
                            email = match.group()
                            if "google.com" not in email.lower() and "gstatic" not in email.lower():
                                data["email"] = email
                                break
            except Exception:
                pass

        if not data["email"]:
            try:
                page_text = page.inner_text('body')
                matches = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', page_text)
                for email in matches:
                    if ("google.com" not in email.lower() and
                        "gstatic" not in email.lower() and
                        "example.com" not in email.lower()):
                        data["email"] = email
                        break
            except Exception:
                pass

        # Website
        el = page.query_selector('a[data-item-id="authority"]')
        if el:
            data["website"] = el.get_attribute("href") or ""
        if not data["website"]:
            el = page.query_selector('button[data-tooltip="Open website"]')
            if el:
                data["website"] = el.inner_text().strip()

        return data

    except Exception as e:
        print(f"    ⚠ Error on #{index}: {e}")
        return None


# ═══════════════════════════════════════════════════════════════
#  SINGLE QUERY RUNNER
# ═══════════════════════════════════════════════════════════════

def scrape_query(page, query: str, max_results: int = 60) -> list[dict]:
    """Run a single query and return the list of records (no-website leads only)."""
    search_url = f"https://www.google.com/maps/search/{quote_plus(query)}"

    print(f"\n  ⏳ Loading: {search_url}")
    try:
        page.goto(search_url, wait_until="load", timeout=60000)
    except Exception as e:
        print(f"  ❌ Failed to load search: {e}")
        return []
    time.sleep(6)

    # Handle consent popup
    for sel in [
        'button:has-text("Accept all")',
        'button:has-text("Accept")',
        'button:has-text("I agree")',
        'form[action*="consent"] button',
    ]:
        try:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                btn.click()
                print("  ✓ Accepted consent")
                time.sleep(4)
                page.goto(search_url, wait_until="load", timeout=60000)
                time.sleep(6)
                break
        except Exception:
            continue

    # Wait for feed
    try:
        page.wait_for_selector('div[role="feed"]', timeout=30000)
    except PWTimeout:
        print("  ⚠ No results feed found — skipping this query")
        return []

    # Scroll & collect URLs
    scroll_results(page, max_results=max_results)
    place_urls = collect_place_urls(page, max_results=max_results)
    total = len(place_urls)
    print(f"  📋 Found {total} places. Visiting each…")

    if total == 0:
        return []

    # Visit each place
    records = []
    skipped_website = 0

    for i, url in enumerate(place_urls):
        print(f"    [{i + 1}/{total}] ", end="", flush=True)
        data = extract_from_place_page(page, url, len(records) + 1)

        if data and data["name"]:
            if data["website"]:
                skipped_website += 1
                print(f"⊘ {data['name']} (has website)")
            else:
                email_str = f" | ✉ {data['email']}" if data['email'] else ""
                records.append(data)
                print(f"✓ {data['name']} | ☎ {data['phone'] or '—'}{email_str}")
        else:
            print("✗ could not extract")

    # Re-number
    for idx, rec in enumerate(records, 1):
        rec["sr_no"] = idx

    print(f"  ── Kept {len(records)} leads (skipped {skipped_website} with websites)")
    return records


# ═══════════════════════════════════════════════════════════════
#  EXCEL EXPORT — COMBINED FILE
# ═══════════════════════════════════════════════════════════════

HEADERS = ["Sr No", "Name", "Lead Score", "Pitch", "Category", "Rating", "Reviews", "Phone", "Email", "Address", "Query"]
KEYS    = ["sr_no", "name", "lead_score", "pitch", "category", "rating", "total_reviews", "phone", "email", "address", "query"]


def style_sheet(ws, records):
    """Apply professional styling to a worksheet."""
    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for r, rec in enumerate(records, 2):
        for c, k in enumerate(KEYS, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(k, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for c in range(1, len(HEADERS) + 1):
        mx = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(records) + 2))
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 50)

    ws.freeze_panes = "A2"


def export_combined_excel(all_results: dict[str, list[dict]], output_dir: str) -> str:
    """
    Export all results into one Excel file.
      - One sheet per query
      - One "All Results" sheet combining everything
    """
    wb = Workbook()

    # ── Sheet 1: All Results Combined ──
    ws_all = wb.active
    ws_all.title = "All Results"
    combined = []
    global_sr = 1
    for query, records in all_results.items():
        for rec in records:
            row = dict(rec)
            row["query"] = query
            row["sr_no"] = global_sr
            combined.append(row)
            global_sr += 1

    style_sheet(ws_all, combined)

    # ── One sheet per query ──
    for query, records in all_results.items():
        sheet_name = sanitize_filename(query)[:31]  # Excel max sheet name = 31 chars

        # Avoid duplicate sheet names
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:27] + f"_{len(existing)}"

        ws = wb.create_sheet(title=sheet_name)
        for rec in records:
            rec["query"] = query
        style_sheet(ws, records)

    # ── Summary sheet ──
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["Query", "Leads Found", "Status"])

    sfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    sfill = PatternFill(start_color="34a853", end_color="34a853", fill_type="solid")
    for c in range(1, 4):
        cell = ws_summary.cell(row=1, column=c)
        cell.font = sfont
        cell.fill = sfill
        cell.alignment = Alignment(horizontal="center")

    for i, (query, records) in enumerate(all_results.items(), 2):
        ws_summary.cell(row=i, column=1, value=query)
        ws_summary.cell(row=i, column=2, value=len(records))
        ws_summary.cell(row=i, column=3, value="✓" if records else "No leads")

    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15

    # ── Total row ──
    total_row = len(all_results) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(combined)).font = Font(bold=True)

    # Save
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"batch_results_{ts}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════
#  MAIN — BATCH RUNNER
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Batch Google Maps Scraper")
    parser.add_argument("--file", default="queries.txt", help="Path to queries file (one per line)")
    parser.add_argument("--max", type=int, default=60, help="Max results per query (default: 60)")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    args = parser.parse_args()

    # ── Read queries ──
    queries_file = args.file
    if not os.path.exists(queries_file):
        print(f"❌ Queries file not found: {queries_file}")
        print(f"   Create it with one query per line, e.g.:")
        print(f"     interior designers Mumbai")
        print(f"     salons in Andheri")
        sys.exit(1)

    with open(queries_file, "r", encoding="utf-8") as f:
        queries = [line.strip() for line in f if line.strip() and not line.startswith("#")]

    if not queries:
        print("❌ No queries found in the file.")
        sys.exit(1)

    # ── Create output directory ──
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(output_dir, exist_ok=True)

    # ── Print banner ──
    print("\n╔════════════════════════════════════════════════════════╗")
    print("║   Batch Google Maps Scraper — No-Website Leads Only   ║")
    print("╠════════════════════════════════════════════════════════╣")
    print(f"║   Queries file : {queries_file:<37s} ║")
    print(f"║   Total queries: {len(queries):<37d} ║")
    print(f"║   Max/query    : {args.max:<37d} ║")
    print(f"║   Headless     : {'Yes' if args.headless else 'No (visible browser)':<37s} ║")
    print("╚════════════════════════════════════════════════════════╝")

    print("\n📋 Queries to scrape:")
    for i, q in enumerate(queries, 1):
        print(f"   {i:>2}. {q}")

    # ── Launch browser ONCE, reuse for all queries ──
    all_results: dict[str, list[dict]] = {}
    start_time = time.time()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=100)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = ctx.new_page()

        for i, query in enumerate(queries, 1):
            print(f"\n{'━' * 58}")
            print(f"  🔍 [{i}/{len(queries)}] \"{query}\"")
            print(f"{'━' * 58}")

            try:
                records = scrape_query(page, query, max_results=args.max)
                all_results[query] = records
            except Exception as e:
                print(f"  ❌ Query failed: {e}")
                all_results[query] = []

            # Small delay between queries to avoid rate limiting
            if i < len(queries):
                print(f"\n  ⏸ Waiting 5s before next query…")
                time.sleep(5)

        browser.close()

    elapsed = time.time() - start_time
    elapsed_str = time.strftime("%H:%M:%S", time.gmtime(elapsed))

    # ── AI Enrichment ──
    print("\n" + "═" * 58)
    print("  🧠 Starting AWS Bedrock AI Enrichment")
    print("═" * 58)
    for query in all_results:
        if all_results[query]:
            all_results[query] = enrich_records(all_results[query])

    # ── Export combined Excel ──
    total_leads = sum(len(r) for r in all_results.values())

    print(f"\n{'═' * 58}")
    print(f"  📊 BATCH COMPLETE")
    print(f"{'═' * 58}")
    print(f"  Queries run   : {len(queries)}")
    print(f"  Total leads   : {total_leads}")
    print(f"  Time elapsed  : {elapsed_str}")
    print(f"{'═' * 58}")

    if total_leads > 0:
        filepath = export_combined_excel(all_results, output_dir)
        print(f"\n  ✅ Excel saved to:\n     {filepath}")
    else:
        print("\n  ⚠ No leads found across all queries.")

    # Per-query breakdown
    print(f"\n  📋 Breakdown:")
    for query, records in all_results.items():
        status = f"✓ {len(records)} leads" if records else "✗ 0 leads"
        print(f"     • {query}: {status}")

    print()


if __name__ == "__main__":
    main()
