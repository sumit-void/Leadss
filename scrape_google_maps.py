"""
Google Maps Scraper — Playwright + Excel Export
  - Only keeps businesses WITHOUT a website
  - Extracts email if available
  - Max 60 results
  - Exports to Excel
"""

import re
import time
import os
from datetime import datetime
from urllib.parse import quote_plus

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()[:80]


def scroll_results(page, max_results: int = 60):
    """Scroll the results feed to load listings."""
    feed = page.query_selector('div[role="feed"]')
    if not feed:
        return 0

    last_count = 0
    stale = 0

    while stale < 10:
        items = page.query_selector_all('div[role="feed"] > div > div > a')
        count = len(items)

        if count >= max_results:
            print(f"  ✓ Loaded {count} listings")
            break

        if count == last_count:
            stale += 1
        else:
            stale = 0
            last_count = count
            print(f"  … {count} listings loaded")

        feed.evaluate("el => el.scrollTop = el.scrollHeight")
        time.sleep(1.8)

    return min(len(page.query_selector_all('div[role="feed"] > div > div > a')), max_results)


def collect_place_urls(page, max_results: int = 60) -> list[str]:
    """Grab all href URLs from listing cards in the feed."""
    links = page.query_selector_all('div[role="feed"] > div > div > a')
    urls = []
    for link in links[:max_results]:
        href = link.get_attribute("href")
        if href and "/maps/place/" in href:
            urls.append(href)
    return urls


def extract_from_place_page(page, url: str, index: int) -> dict | None:
    """Navigate to a place page and extract business details."""
    try:
        page.goto(url, wait_until="load", timeout=30000)
        time.sleep(3)

        # Wait for the business name
        try:
            page.wait_for_selector('h1.DUwDvf, h1.fontHeadlineLarge', timeout=10000)
        except PWTimeout:
            try:
                page.wait_for_selector('h1', timeout=5000)
            except PWTimeout:
                return None

        # ── SCROLL the detail panel to reveal ALL info fields ──
        # Email, hours, etc. are often below the fold in the scrollable panel
        detail_panels = page.query_selector_all('div[role="main"]')
        for panel in detail_panels:
            try:
                for _ in range(5):
                    panel.evaluate("el => el.scrollTop += 500")
                    time.sleep(0.3)
                panel.evaluate("el => el.scrollTop = 0")
                time.sleep(0.3)
            except Exception:
                pass

        # Also try the scrollable info container
        try:
            page.evaluate("""
                const scrollable = document.querySelector('[class*="m6QErb"][class*="DxyBCb"]');
                if (scrollable) {
                    for (let i = 0; i < 6; i++) {
                        scrollable.scrollTop += 400;
                    }
                    scrollable.scrollTop = 0;
                }
            """)
            time.sleep(0.5)
        except Exception:
            pass

        data = {
            "sr_no": index,
            "name": "",
            "address": "",
            "phone": "",
            "email": "",
            "website": "",
            "rating": "",
            "total_reviews": "",
            "category": "",
        }

        # Name
        for sel in ['h1.DUwDvf', 'h1.fontHeadlineLarge', 'h1']:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                if txt and txt.lower() != "results":
                    data["name"] = txt
                    break

        if not data["name"]:
            return None

        # Rating
        el = page.query_selector('div.F7nice span[aria-hidden="true"]')
        if el:
            data["rating"] = el.inner_text().strip()

        # Reviews
        el = page.query_selector('div.F7nice span[aria-label*="review"]')
        if not el:
            el = page.query_selector('div.F7nice span:nth-child(2)')
        if el:
            data["total_reviews"] = el.inner_text().strip().replace("(", "").replace(")", "").replace(",", "")

        # Category
        el = page.query_selector('button[jsaction*="category"]')
        if el:
            data["category"] = el.inner_text().strip()

        # Address
        for sel in ['button[data-item-id="address"]', 'button[data-tooltip="Copy address"]']:
            el = page.query_selector(sel)
            if el:
                data["address"] = el.inner_text().strip()
                break

        # Phone
        for sel in ['button[data-item-id*="phone:tel"]', 'button[data-tooltip="Copy phone number"]']:
            el = page.query_selector(sel)
            if el:
                data["phone"] = el.inner_text().strip()
                break

        # ── EMAIL — 3-layer extraction ──

        # Layer 1: Google Maps native email elements
        email_selectors = [
            'a[data-item-id*="email"]',
            'button[data-item-id*="email"]',
            'a[href^="mailto:"]',
            'a[data-tooltip="Send email"]',
            'button[data-tooltip="Send email"]',
            '[data-item-id*="email"]',
            'a[aria-label*="email"]',
            'a[aria-label*="Email"]',
            'button[aria-label*="email"]',
            'button[aria-label*="Email"]',
        ]
        for sel in email_selectors:
            el = page.query_selector(sel)
            if el:
                href = el.get_attribute("href") or ""
                if href.startswith("mailto:"):
                    data["email"] = href.replace("mailto:", "").split("?")[0].strip()
                else:
                    aria = el.get_attribute("aria-label") or ""
                    txt = el.inner_text().strip()
                    email_in_aria = re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', aria)
                    if email_in_aria:
                        data["email"] = email_in_aria.group()
                    elif "@" in txt:
                        data["email"] = txt
                if data["email"]:
                    break

        # Layer 2: Scan info section buttons/links for email text
        if not data["email"]:
            try:
                info_elements = page.query_selector_all('div[role="main"] button, div[role="main"] a')
                for elem in info_elements:
                    txt = elem.inner_text().strip()
                    if "@" in txt and "." in txt:
                        match = re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', txt)
                        if match:
                            email = match.group()
                            if "google.com" not in email.lower() and "gstatic" not in email.lower():
                                data["email"] = email
                                break
            except Exception:
                pass

        # Layer 3: Full page text regex scan
        if not data["email"]:
            try:
                page_text = page.inner_text('body')
                matches = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', page_text)
                for email in matches:
                    if ("google.com" not in email.lower() and
                        "gstatic" not in email.lower() and
                        "example.com" not in email.lower()):
                        data["email"] = email
                        break
            except Exception:
                pass

        # Website
        el = page.query_selector('a[data-item-id="authority"]')
        if el:
            data["website"] = el.get_attribute("href") or ""
        if not data["website"]:
            el = page.query_selector('button[data-tooltip="Open website"]')
            if el:
                data["website"] = el.inner_text().strip()

        return data

    except Exception as e:
        print(f"  ⚠ Error on #{index}: {e}")
        return None


def export_to_excel(records: list[dict], query: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Google Maps Data"

    headers = ["Sr No", "Name", "Category", "Rating", "Reviews", "Phone", "Email", "Address"]
    keys = ["sr_no", "name", "category", "rating", "total_reviews", "phone", "email", "address"]

    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for r, rec in enumerate(records, 2):
        for c, k in enumerate(keys, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(k, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for c in range(1, len(headers) + 1):
        mx = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(records) + 2))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(mx + 4, 50)

    ws.freeze_panes = "A2"

    safe = sanitize_filename(query)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"gmaps_{safe}_{ts}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    return filepath


def main():
    print("\n╔══════════════════════════════════════════════════╗")
    print("║  Google Maps Scraper — No-Website Leads Only    ║")
    print("╚══════════════════════════════════════════════════╝\n")

    query = input("🔍 Search query (e.g. 'interior designers Mumbai'): ").strip()
    if not query:
        print("No query. Exiting.")
        return

    max_results = 60
    custom = input(f"📊 Max results [{max_results}]: ").strip()
    if custom.isdigit() and int(custom) > 0:
        max_results = min(int(custom), 120)

    search_url = f"https://www.google.com/maps/search/{quote_plus(query)}"

    print(f"\n🚀 Launching browser for: '{query}' (max {max_results})")
    print(f"   URL: {search_url}\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )
        page = ctx.new_page()

        # ── Step 1: Load search results ──
        print("  ⏳ Loading search results…")
        page.goto(search_url, wait_until="load", timeout=60000)
        time.sleep(6)

        # Handle consent
        for sel in [
            'button:has-text("Accept all")',
            'button:has-text("Accept")',
            'button:has-text("I agree")',
            'form[action*="consent"] button',
        ]:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click()
                    print("  ✓ Accepted consent")
                    time.sleep(4)
                    page.goto(search_url, wait_until="load", timeout=60000)
                    time.sleep(6)
                    break
            except Exception:
                continue

        # Wait for feed
        try:
            page.wait_for_selector('div[role="feed"]', timeout=30000)
            print("  ✓ Results feed loaded")
        except PWTimeout:
            print("⚠ No results feed. Saving debug screenshot.")
            page.screenshot(path="debug_gmaps.png")
            browser.close()
            return

        # ── Step 2: Scroll & collect all place URLs ──
        scroll_results(page, max_results=max_results)
        place_urls = collect_place_urls(page, max_results=max_results)
        total = len(place_urls)
        print(f"\n📋 Collected {total} place URLs. Visiting each one…\n")

        if total == 0:
            print("⚠ No place URLs found.")
            browser.close()
            return

        # ── Step 3: Visit each place page & extract ──
        records = []
        skipped_website = 0
        skipped_nodata = 0

        for i, url in enumerate(place_urls):
            print(f"  [{i + 1}/{total}] ", end="", flush=True)

            data = extract_from_place_page(page, url, len(records) + 1)

            if data and data["name"]:
                if data["website"]:
                    skipped_website += 1
                    print(f"⊘ {data['name']} (has website — skipped)")
                else:
                    email_str = f" | ✉ {data['email']}" if data['email'] else ""
                    records.append(data)
                    print(f"✓ {data['name']} | ☎ {data['phone'] or '—'} | ⭐ {data['rating'] or '—'}{email_str}")
            else:
                skipped_nodata += 1
                print("✗ could not extract")

        browser.close()

    # Re-number
    for idx, rec in enumerate(records, 1):
        rec["sr_no"] = idx

    print(f"\n{'─' * 55}")
    print(f"  Total place URLs:      {total}")
    print(f"  Skipped (has website): {skipped_website}")
    print(f"  Skipped (no data):     {skipped_nodata}")
    print(f"  ✅ Kept (no website):   {len(records)}")
    print(f"{'─' * 55}")

    if records:
        filepath = export_to_excel(records, query)
        print(f"\n✅ Exported {len(records)} records to:\n   {filepath}\n")
    else:
        print("\n⚠ All listings had websites or couldn't be extracted.\n")


if __name__ == "__main__":
    main()
