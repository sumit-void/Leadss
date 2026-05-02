"""
JustDial Scraper — Playwright + Excel Export
  - Only keeps businesses WITHOUT a website
  - Max 60 results
  - Exports to Excel
"""

import re
import time
import os
from datetime import datetime
from urllib.parse import quote

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()[:80]


def close_popups(page):
    """Dismiss any JustDial popups, modals, login prompts."""
    popup_selectors = [
        'span.close_btn',
        'button.close',
        'span#best_deal_close',
        'div.modal-close',
        'button[aria-label="Close"]',
        'span[class*="cross"]',
        'a[title="Close"]',
        '#best_deal_popup span',
        'div[class*="popup"] span[class*="close"]',
        'div[class*="modal"] button[class*="close"]',
        'div[class*="overlay"] span[class*="close"]',
    ]
    for sel in popup_selectors:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible():
                el.click()
                time.sleep(0.5)
        except Exception:
            continue


def scroll_and_load(page, max_results: int = 60):
    """Scroll the page and click 'Load More' to get more listings."""
    last_height = 0
    stale = 0

    while stale < 8:
        # Scroll to bottom
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(2)

        close_popups(page)

        # Click any "Load More" / "Show More" buttons
        for sel in [
            'a:has-text("Load More")',
            'button:has-text("Load More")',
            'span:has-text("Load More")',
            'a:has-text("Show More")',
            'button:has-text("Show More")',
            'a[class*="loadMore"]',
            'div[class*="loadMore"]',
        ]:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click()
                    print("  ↻ Clicked Load More")
                    time.sleep(3)
                    break
            except Exception:
                continue

        new_height = page.evaluate("document.body.scrollHeight")
        if new_height == last_height:
            stale += 1
        else:
            stale = 0
            last_height = new_height


def collect_listing_urls(page, city: str, max_results: int = 60) -> list[str]:
    """Collect all individual business page URLs from the search results."""
    urls = []
    seen = set()

    city_lower = city.lower()

    # Grab ALL <a> tags on the page
    all_links = page.query_selector_all('a[href]')

    for link in all_links:
        try:
            href = link.get_attribute("href") or ""

            # JustDial detail pages match: /CityName/BusinessName-.../0XXXXX...
            # They contain the city name and end with a numeric ID
            if not href:
                continue

            # Normalize
            if href.startswith("/"):
                href = "https://www.justdial.com" + href

            # Must be a justdial detail page URL
            if "justdial.com" not in href:
                continue
            if f"/{city_lower}/" not in href.lower() and f"/{city}/" not in href:
                continue

            # Skip non-detail pages (search pages, category pages, etc.)
            # Detail pages have a long numeric suffix like /011PXX...
            # or contain specific path segments
            path = href.split("justdial.com")[-1] if "justdial.com" in href else ""
            segments = [s for s in path.split("/") if s]

            # Need at least city + business-name segments
            if len(segments) < 2:
                continue

            # Skip known non-detail patterns
            skip_patterns = [
                "/nct-", "/page-", "/sort-", "/rating-",
                "login", "signup", "advertise", "feedback",
                "privacy", "terms", "about", "contact",
                "cms/", "support/",
            ]
            if any(pat in href.lower() for pat in skip_patterns):
                continue

            # The business slug should be the 2nd segment and contain hyphens (multi-word)
            # and the URL should not be the search results page itself
            biz_slug = segments[1] if len(segments) > 1 else ""

            # JustDial detail pages typically have a numeric ID at the end
            # or the slug contains enough info
            last_segment = segments[-1]
            has_numeric_id = bool(re.search(r'\d{8,}', last_segment))

            # Also accept URLs where the slug has a phone-number-like pattern
            has_phone_pattern = bool(re.search(r'\d{10,}', href))

            if has_numeric_id or has_phone_pattern:
                if href not in seen:
                    seen.add(href)
                    urls.append(href)

            if len(urls) >= max_results:
                break

        except Exception:
            continue

    return urls[:max_results]


def extract_from_detail_page(page, url: str, index: int) -> dict | None:
    """Navigate to a JustDial business detail page and extract info."""
    try:
        page.goto(url, wait_until="load", timeout=30000)
        time.sleep(3)

        close_popups(page)

        data = {
            "sr_no": index,
            "name": "",
            "address": "",
            "phone": "",
            "website": "",
            "rating": "",
            "total_reviews": "",
            "category": "",
        }

        # ── Name ──
        for sel in [
            'h1', 'span.fn', 'h2.company-name',
            'h1[class*="name"]', 'span[class*="title"]',
            'p.fn', 'div.fn',
        ]:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                if txt and len(txt) > 1:
                    data["name"] = txt
                    break

        if not data["name"]:
            return None

        # ── Rating ──
        for sel in [
            'span.green-box', 'span[class*="rating"]',
            'span.total-rating', 'div[class*="rating"] span',
            'span.star_m', 'i.star',
        ]:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                if txt and any(c.isdigit() for c in txt):
                    data["rating"] = txt
                    break

        # ── Reviews ──
        for sel in [
            'span.rt_count', 'a[class*="review"] span',
            'span[class*="votes"]', 'span[class*="review"]',
            'a[class*="rating"]', 'span.total_r',
        ]:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                match = re.search(r'[\d,]+', txt)
                if match:
                    data["total_reviews"] = match.group()
                    break

        # ── Address ──
        for sel in [
            'span[class*="address"]', 'p[class*="address"]',
            'span.lng_br_addr', 'div[class*="address"]',
            'span.cont_fl_addr', 'span.mcelavs',
            'p.address', 'span.adr',
        ]:
            el = page.query_selector(sel)
            if el:
                txt = el.inner_text().strip()
                if txt and len(txt) > 5:
                    data["address"] = txt
                    break

        # If address not found, try to get from the page text
        if not data["address"]:
            addr_el = page.query_selector('section[class*="address"], div[class*="location"]')
            if addr_el:
                data["address"] = addr_el.inner_text().strip()[:200]

        # ── Category ──
        for sel in [
            'span.lng_br_sname', 'span[class*="category"]',
            'p[class*="category"]', 'span.cat_name',
            'div[class*="category"]', 'a[class*="tag"]',
        ]:
            el = page.query_selector(sel)
            if el:
                data["category"] = el.inner_text().strip()
                break

        # ── Phone ──
        # Method 1: tel: links
        for sel in ['a[href^="tel:"]']:
            el = page.query_selector(sel)
            if el:
                href = el.get_attribute("href") or ""
                if href.startswith("tel:"):
                    phone = href.replace("tel:", "").strip()
                    if len(phone) >= 8:
                        data["phone"] = phone
                        break

        # Method 2: Look for phone display elements
        if not data["phone"]:
            for sel in [
                'span.mobilesv', 'span[class*="phone"]',
                'p[class*="phone"]', 'span[class*="contact"]',
                'a[class*="phone"]', 'div[class*="phone"]',
            ]:
                el = page.query_selector(sel)
                if el:
                    txt = el.inner_text().strip()
                    match = re.search(r'[\d\s\-+]{8,}', txt)
                    if match:
                        data["phone"] = match.group().strip()
                        break

        # Method 3: Click "Show Number" button
        if not data["phone"]:
            for sel in [
                'button:has-text("Show Number")',
                'a:has-text("Show Number")',
                'span:has-text("Show Number")',
                'p:has-text("Show Number")',
                'button:has-text("Contact")',
            ]:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        btn.click()
                        time.sleep(2)
                        # Try to get the revealed number
                        tel = page.query_selector('a[href^="tel:"]')
                        if tel:
                            href = tel.get_attribute("href") or ""
                            if href.startswith("tel:"):
                                data["phone"] = href.replace("tel:", "").strip()
                                break
                except Exception:
                    continue

        # ── Website ──
        # Look for external website links (not justdial links)
        for sel in [
            'a[class*="website"]', 'a[data-type="website"]',
            'a[title="Website"]', 'a:has-text("Website")',
            'a[class*="web"]',
        ]:
            el = page.query_selector(sel)
            if el:
                href = el.get_attribute("href") or ""
                if href and "justdial" not in href.lower() and href.startswith("http"):
                    data["website"] = href
                    break

        # Also check all links for external websites
        if not data["website"]:
            all_a = page.query_selector_all('a[href^="http"]')
            for a in all_a:
                try:
                    href = a.get_attribute("href") or ""
                    if (href and
                        "justdial" not in href.lower() and
                        "google" not in href.lower() and
                        "facebook.com" not in href.lower() and
                        "twitter.com" not in href.lower() and
                        "instagram.com" not in href.lower() and
                        "youtube.com" not in href.lower() and
                        "linkedin.com" not in href.lower() and
                        "play.google" not in href.lower() and
                        "apps.apple" not in href.lower() and
                        "whatsapp" not in href.lower() and
                        not href.endswith(".png") and
                        not href.endswith(".jpg")):
                        data["website"] = href
                        break
                except Exception:
                    continue

        return data

    except Exception as e:
        print(f"  ⚠ Error on #{index}: {e}")
        return None


def export_to_excel(records: list[dict], query: str, city: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "JustDial Data"

    headers = ["Sr No", "Name", "Category", "Rating", "Reviews", "Phone", "Address"]
    keys = ["sr_no", "name", "category", "rating", "total_reviews", "phone", "address"]

    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2ecc40", end_color="2ecc40", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for r, rec in enumerate(records, 2):
        for c, k in enumerate(keys, 1):
            cell = ws.cell(row=r, column=c, value=rec.get(k, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for c in range(1, len(headers) + 1):
        mx = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(records) + 2))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(mx + 4, 50)

    ws.freeze_panes = "A2"

    safe = sanitize_filename(f"{city}_{query}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"justdial_{safe}_{ts}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    return filepath


def main():
    print("\n╔══════════════════════════════════════════════════╗")
    print("║   JustDial Scraper — No-Website Leads Only      ║")
    print("╚══════════════════════════════════════════════════╝\n")

    city = input("🏙️  City (e.g. 'Mumbai'): ").strip()
    if not city:
        print("No city. Exiting.")
        return

    query = input("🔍 Search query (e.g. 'interior designers'): ").strip()
    if not query:
        print("No query. Exiting.")
        return

    max_results = 60
    custom = input(f"📊 Max results [{max_results}]: ").strip()
    if custom.isdigit() and int(custom) > 0:
        max_results = min(int(custom), 120)

    city_slug = city.strip().capitalize()
    query_slug = quote(query.strip())
    search_url = f"https://www.justdial.com/{city_slug}/{query_slug}/nct-10000000"

    print(f"\n🚀 Launching browser for: '{query}' in {city} (max {max_results})")
    print(f"   URL: {search_url}\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=100)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        # ── Step 1: Load search results ──────────────────────────────
        print("  ⏳ Loading search results…")
        try:
            page.goto(search_url, wait_until="load", timeout=45000)
        except PWTimeout:
            print("  ⚠ Page load timed out, continuing…")

        time.sleep(5)
        close_popups(page)
        time.sleep(1)

        # ── Step 2: Scroll to load more results ─────────────────────
        print("  ⏳ Scrolling to load listings…")
        scroll_and_load(page, max_results=max_results)

        # ── Step 3: Collect all listing detail URLs ──────────────────
        print("  ⏳ Collecting listing URLs…")
        listing_urls = collect_listing_urls(page, city_slug, max_results=max_results)
        total = len(listing_urls)
        print(f"\n📋 Collected {total} listing URLs. Visiting each one…\n")

        if total == 0:
            print("  ⚠ No listing URLs found. Trying alternate approach…\n")

            # Fallback: Try to extract data directly from the search page
            # This handles the case where JustDial doesn't use standard detail URLs
            page.screenshot(path="debug_justdial.png")
            print("  Screenshot saved to debug_justdial.png for debugging.")

            # Try to extract names and info from visible cards
            fallback_records = []
            cards = page.query_selector_all('li, div[class*="result"], div[class*="store"], div[class*="card"]')

            for i, card in enumerate(cards[:max_results]):
                try:
                    # Try to find a name
                    name_el = None
                    for sel in ['h2', 'h3', 'a[class*="title"]', 'span[class*="name"]', 'p[class*="name"]']:
                        name_el = card.query_selector(sel)
                        if name_el:
                            break
                    if not name_el:
                        continue

                    name = name_el.inner_text().strip()
                    if not name or len(name) < 2:
                        continue

                    # Basic extraction
                    rec = {
                        "sr_no": len(fallback_records) + 1,
                        "name": name,
                        "category": "",
                        "rating": "",
                        "total_reviews": "",
                        "phone": "",
                        "address": "",
                        "website": "",
                    }

                    # Try rating
                    rating_el = card.query_selector('span[class*="rating"], span[class*="green"]')
                    if rating_el:
                        rec["rating"] = rating_el.inner_text().strip()

                    # Try address
                    addr_el = card.query_selector('span[class*="addr"], p[class*="addr"], span[class*="location"]')
                    if addr_el:
                        rec["address"] = addr_el.inner_text().strip()

                    # Try phone
                    tel_el = card.query_selector('a[href^="tel:"]')
                    if tel_el:
                        rec["phone"] = (tel_el.get_attribute("href") or "").replace("tel:", "").strip()

                    if not rec.get("website"):
                        fallback_records.append(rec)
                        print(f"  [{len(fallback_records)}] ✓ {name}")

                except Exception:
                    continue

            if fallback_records:
                filepath = export_to_excel(fallback_records, query, city)
                print(f"\n✅ Exported {len(fallback_records)} records (fallback mode):\n   {filepath}\n")
            else:
                print("\n⚠ Could not extract any listings. JustDial's layout may have changed.\n")

            browser.close()
            return

        # ── Step 4: Visit each detail page ───────────────────────────
        records = []
        skipped_website = 0
        skipped_nodata = 0

        for i, url in enumerate(listing_urls):
            print(f"  [{i + 1}/{total}] ", end="", flush=True)

            data = extract_from_detail_page(page, url, len(records) + 1)

            if data and data["name"]:
                if data["website"]:
                    skipped_website += 1
                    print(f"⊘ {data['name']} (has website — skipped)")
                else:
                    records.append(data)
                    print(f"✓ {data['name']} | ☎ {data['phone'] or '—'} | ⭐ {data['rating'] or '—'}")
            else:
                skipped_nodata += 1
                print("✗ could not extract")

        browser.close()

    # Re-number
    for idx, rec in enumerate(records, 1):
        rec["sr_no"] = idx

    print(f"\n{'─' * 55}")
    print(f"  Total listing URLs:    {total}")
    print(f"  Skipped (has website): {skipped_website}")
    print(f"  Skipped (no data):     {skipped_nodata}")
    print(f"  ✅ Kept (no website):   {len(records)}")
    print(f"{'─' * 55}")

    if records:
        filepath = export_to_excel(records, query, city)
        print(f"\n✅ Exported {len(records)} records to:\n   {filepath}\n")
    else:
        print("\n⚠ All listings had websites or couldn't be extracted.\n")


if __name__ == "__main__":
    main()
